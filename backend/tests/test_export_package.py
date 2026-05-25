import csv
import json
import shutil
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.services.export_package import BACKEND_ROOT, EXPORT_ROOT

client = TestClient(app)

RAW_SENTINEL = "PRIVATE RAW TEXT SENTINEL"


def sample_capture_result(max_results: int = 1) -> dict[str, object]:
    raw_jobs = [
        {
            "source": "manual_raw_jobs",
            "source_url": "https://example.test/export-job",
            "raw_text": f"""
Title: Microsoft 365 Support Specialist
Company: Example SaaS
Location: Remote, Spain
Work mode: Remote
English required. {RAW_SENTINEL}
Support Microsoft 365, Entra ID, endpoint troubleshooting, and SaaS workflows.
""",
            "external_id": "export-test-1",
            "capture_notes": ["test fixture"],
        },
    ]
    if max_results > 1:
        raw_jobs.extend(
            [
                {
                    "source": "manual_raw_jobs",
                    "source_url": "https://example.test/manual-review-job",
                    "raw_text": """
Support role.
Company unknown.
Tickets and users.
Need help soon.
""",
                    "external_id": "export-test-2",
                    "capture_notes": ["test fixture"],
                },
                {
                    "source": "manual_raw_jobs",
                    "source_url": "https://example.test/duplicate-job",
                    "raw_text": """
Title: Duplicate Support Specialist
Company: Example SaaS
Location: Remote, Spain
Work mode: Remote
English required. Microsoft 365 support.
""",
                    "external_id": "export-test-3",
                    "capture_notes": ["test fixture"],
                },
            ]
        )

    response = client.post(
        "/api/capture/run",
        json={
            "profile_id": "rafael_default",
            "source": "manual_raw_jobs",
            "max_results": max_results,
            "dry_run": True,
            "raw_jobs": raw_jobs,
        },
    )
    assert response.status_code == 200
    data = response.json()
    if max_results > 1:
        duplicate_result = data["results"][2]
        duplicate_result["duplicate_preview"] = True
        duplicate_result["duplicate_reason"] = "same source_url as saved history item"
    return data


def export_result(
    export_format: str,
    include_raw_text: bool,
    capture_result: dict[str, object] | None = None,
) -> dict[str, object]:
    response = client.post(
        "/api/export/capture-result",
        json={
            "export_format": export_format,
            "include_raw_text": include_raw_text,
            "capture_result": capture_result or sample_capture_result(),
        },
    )
    assert response.status_code == 200
    return response.json()


def export_file_path(response_data: dict[str, object]) -> Path:
    files = response_data["files"]
    assert isinstance(files, list)
    assert len(files) == 1
    path = (BACKEND_ROOT / files[0]).resolve()
    assert EXPORT_ROOT.resolve() in path.parents
    return path


def cleanup_export(response_data: dict[str, object]) -> None:
    path = export_file_path(response_data)
    export_dir = path.parent
    if export_dir.parent == EXPORT_ROOT.resolve() and export_dir.name.startswith("export_"):
        shutil.rmtree(export_dir, ignore_errors=True)


def sheet_column_values(workbook, sheet_name: str, column_name: str) -> list[object]:
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    column_index = headers.index(column_name) + 1
    return [sheet.cell(row=row_index, column=column_index).value for row_index in range(2, sheet.max_row + 1)]


def test_export_json_from_sample_capture_result() -> None:
    data = export_result("json", include_raw_text=True)
    try:
        path = export_file_path(data)
        payload = json.loads(path.read_text(encoding="utf-8"))

        assert data["status"] == "completed"
        assert path.suffix == ".json"
        assert payload["profile_id"] == "rafael_default"
        assert payload["results"][0]["decision"]["decision"] in {"Apply", "Maybe", "Manual Review", "Discard"}
        assert RAW_SENTINEL in payload["results"][0]["raw_job"]["raw_text"]
    finally:
        cleanup_export(data)


def test_export_csv_from_sample_capture_result() -> None:
    data = export_result("csv", include_raw_text=False)
    try:
        path = export_file_path(data)
        rows = list(csv.DictReader(path.read_text(encoding="utf-8").splitlines()))

        assert path.suffix == ".csv"
        assert len(rows) == 1
        assert rows[0]["title"] == "Microsoft 365 Support Specialist"
        assert rows[0]["source_url"] == "https://example.test/export-job"
        assert "raw_text" not in rows[0]
        assert any("Raw job text was excluded" in warning for warning in data["warnings"])
    finally:
        cleanup_export(data)


def test_export_xlsx_from_sample_capture_result() -> None:
    data = export_result("xlsx", include_raw_text=False)
    try:
        path = export_file_path(data)
        workbook = load_workbook(path)
        sheet = workbook["All Reviewed Jobs"]
        headers = [cell.value for cell in sheet[1]]
        first_row = [cell.value for cell in sheet[2]]

        assert path.suffix == ".xlsx"
        assert "decision" in headers
        assert "status" in headers
        assert "raw_text" not in headers
        assert "Microsoft 365 Support Specialist" in first_row
        assert "Summary" in workbook.sheetnames
    finally:
        cleanup_export(data)


def test_export_xlsx_contains_workflow_sheets_and_summary() -> None:
    data = export_result("xlsx", include_raw_text=False, capture_result=sample_capture_result(max_results=3))
    try:
        path = export_file_path(data)
        workbook = load_workbook(path)

        expected = {
            "Summary",
            "All Reviewed Jobs",
            "Apply Today",
            "Manual Review",
            "Waiting Follow Up",
            "Duplicates Reviewed",
            "Decision Explanations",
            "Capture Diagnostics",
        }
        assert expected.issubset(set(workbook.sheetnames))

        summary_values = {
            workbook["Summary"].cell(row=index, column=1).value: workbook["Summary"].cell(row=index, column=2).value
            for index in range(2, workbook["Summary"].max_row + 1)
        }
        assert summary_values["Total jobs"] == 3
        assert summary_values["Profile ID"] == "rafael_default"
        assert summary_values["Duplicate / already reviewed"] == 1
    finally:
        cleanup_export(data)


def test_export_xlsx_queue_sheets_filter_rows() -> None:
    data = export_result("xlsx", include_raw_text=False, capture_result=sample_capture_result(max_results=3))
    try:
        path = export_file_path(data)
        workbook = load_workbook(path)

        apply_titles = sheet_column_values(workbook, "Apply Today", "title")
        manual_titles = sheet_column_values(workbook, "Manual Review", "title")
        duplicate_titles = sheet_column_values(workbook, "Duplicates Reviewed", "title")

        assert "Microsoft 365 Support Specialist" in apply_titles
        assert len(manual_titles) >= 1
        assert any(title in {None, "", "Duplicate Support Specialist"} for title in manual_titles)
        assert "Duplicate Support Specialist" in duplicate_titles
    finally:
        cleanup_export(data)


def test_export_xlsx_decision_explanations_sheet_has_rules() -> None:
    data = export_result("xlsx", include_raw_text=False, capture_result=sample_capture_result(max_results=3))
    try:
        path = export_file_path(data)
        workbook = load_workbook(path)
        headers = [cell.value for cell in workbook["Decision Explanations"][1]]

        assert "triggered_rules" in headers
        assert "matched_positive_keywords" in headers
        assert "missing_information" in headers
    finally:
        cleanup_export(data)


def test_include_raw_text_false_excludes_raw_text_from_json() -> None:
    data = export_result("json", include_raw_text=False)
    try:
        path = export_file_path(data)
        payload = json.loads(path.read_text(encoding="utf-8"))

        assert payload["results"][0]["raw_job"]["raw_text"] == ""
        assert RAW_SENTINEL not in path.read_text(encoding="utf-8")
    finally:
        cleanup_export(data)


def test_include_raw_text_true_includes_raw_text_in_csv() -> None:
    data = export_result("csv", include_raw_text=True)
    try:
        path = export_file_path(data)
        rows = list(csv.DictReader(path.read_text(encoding="utf-8").splitlines()))

        assert rows[0]["raw_text"]
        assert RAW_SENTINEL in rows[0]["raw_text"]
    finally:
        cleanup_export(data)


def test_export_output_path_stays_under_backend_data_exports() -> None:
    data = export_result("json", include_raw_text=False)
    try:
        path = export_file_path(data)
        relative = path.relative_to(BACKEND_ROOT).as_posix()

        assert relative.startswith("data/exports/export_")
        assert path.exists()
    finally:
        cleanup_export(data)


def test_invalid_export_format_returns_controlled_error() -> None:
    response = client.post(
        "/api/export/capture-result",
        json={
            "export_format": "pdf",
            "include_raw_text": False,
            "capture_result": sample_capture_result(),
        },
    )

    assert response.status_code == 422


def test_generated_exports_are_ignored_by_git() -> None:
    gitignore = (BACKEND_ROOT.parent / ".gitignore").read_text(encoding="utf-8")

    assert "backend/data/" in gitignore
    assert "*.xlsx" in gitignore
