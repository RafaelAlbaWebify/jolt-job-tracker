import csv
import json
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.models import (
    CaptureJobResult,
    CaptureRunResult,
    ExportCaptureResultResponse,
    ExportFormat,
    HistoryJobEntry,
)

BACKEND_ROOT = Path(__file__).resolve().parents[2]
EXPORT_ROOT = BACKEND_ROOT / "data" / "exports"

BASE_COLUMNS = [
    "status",
    "decision",
    "priority",
    "score",
    "title",
    "company",
    "location",
    "work_mode",
    "profile_id",
    "parser_confidence",
    "capture_confidence",
    "reasons",
    "triggered_rules",
    "warnings",
    "missing_information",
    "matched_positive_keywords",
    "matched_risk_keywords",
    "source_url",
    "created_at",
    "errors",
]

EXPLANATION_COLUMNS = [
    "title",
    "company",
    "decision",
    "triggered_rules",
    "matched_positive_keywords",
    "matched_risk_keywords",
    "hard_discard_reasons",
    "warnings",
    "missing_information",
]


def _model_to_dict(model: Any) -> dict[str, Any]:
    if hasattr(model, "model_dump"):
        return model.model_dump()
    return model.dict()


def _join(values: list[str] | None) -> str:
    return "; ".join(values or [])


def _status_for_result(result: CaptureJobResult) -> str:
    decision = result.decision.decision if result.decision else ""
    if result.duplicate_preview:
        return "Duplicate"
    if decision == "Apply":
        return "Apply Today"
    if decision in {"Manual Review", "Duplicate", "Already Reviewed"}:
        return decision
    if decision == "Discard":
        return "Archived"
    return "New"


def _relative_to_backend(path: Path) -> str:
    return path.resolve().relative_to(BACKEND_ROOT).as_posix()


def _ensure_export_dir(export_id: str) -> Path:
    export_dir = (EXPORT_ROOT / export_id).resolve()
    export_root = EXPORT_ROOT.resolve()
    if export_root not in export_dir.parents and export_dir != export_root:
        raise ValueError("Export path escaped the configured export root.")

    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


def _flatten_result(result: CaptureJobResult, include_raw_text: bool) -> dict[str, Any]:
    job = result.parsed_job
    decision = result.decision
    row: dict[str, Any] = {
        "status": _status_for_result(result),
        "decision": "Duplicate" if result.duplicate_preview else decision.decision if decision else "",
        "priority": decision.priority if decision else "",
        "score": decision.score if decision else "",
        "title": job.title if job else "",
        "company": job.company if job else "",
        "location": job.location if job else "",
        "work_mode": job.work_mode if job else "",
        "profile_id": decision.profile_id if decision else "",
        "parser_confidence": job.parser_confidence if job else "",
        "capture_confidence": "",
        "reasons": _join(decision.reasons if decision else []),
        "triggered_rules": _join(decision.triggered_rules if decision else []),
        "warnings": _join(decision.warnings if decision else []),
        "missing_information": _join(decision.missing_information if decision else []),
        "matched_positive_keywords": _join(decision.matched_positive_keywords if decision else []),
        "matched_risk_keywords": _join(decision.matched_risk_keywords if decision else []),
        "source_url": result.raw_job.source_url or (job.source_url if job else ""),
        "created_at": result.raw_job.captured_at,
        "errors": _join(result.errors),
    }
    if include_raw_text:
        row["raw_text"] = result.raw_job.raw_text
    return row


def _json_payload(capture_result: CaptureRunResult, include_raw_text: bool) -> dict[str, Any]:
    payload = _model_to_dict(capture_result)
    if include_raw_text:
        return payload

    for result in payload["results"]:
        result["raw_job"]["raw_text"] = ""
        if result.get("parsed_job"):
            result["parsed_job"]["description"] = ""
    return payload


def _write_json(path: Path, capture_result: CaptureRunResult, include_raw_text: bool) -> None:
    path.write_text(
        json.dumps(_json_payload(capture_result, include_raw_text), indent=2),
        encoding="utf-8",
    )


def _write_csv(path: Path, rows: list[dict[str, Any]], include_raw_text: bool) -> None:
    columns = BASE_COLUMNS + (["raw_text"] if include_raw_text else [])
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def _style_header(sheet: Worksheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="EAF3F1")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="21313A")
        cell.fill = fill


def _autosize_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        header = column_cells[0]
        if header.column_letter is None:
            continue
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[header.column_letter].width = min(max(max_length + 2, 12), 48)


def _append_table(sheet: Worksheet, columns: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _style_header(sheet)
    _autosize_columns(sheet)

    if "source_url" in columns:
        url_index = columns.index("source_url") + 1
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=url_index)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"


def _summary_rows(capture_result: CaptureRunResult, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return _summary_rows_for_export(
        rows,
        metadata=[
            {"metric": "Export timestamp", "value": datetime.now(UTC).isoformat(timespec="seconds")},
            {"metric": "Run ID", "value": capture_result.run_id},
            {"metric": "Profile ID", "value": capture_result.profile_id},
            {"metric": "Capture mode", "value": capture_result.capture_diagnostics.capture_mode_used},
            {"metric": "Capture confidence", "value": capture_result.capture_diagnostics.capture_confidence},
            {"metric": "Total jobs", "value": len(rows)},
            {"metric": "Parsed jobs", "value": capture_result.parsed_count},
            {"metric": "Classified jobs", "value": capture_result.classified_count},
            {"metric": "Failed jobs", "value": capture_result.failed_count},
        ],
    )


def _summary_rows_for_export(rows: list[dict[str, Any]], metadata: list[dict[str, Any]]) -> list[dict[str, Any]]:
    decision_counts = Counter(row.get("decision", "") or "Unclassified" for row in rows)
    status_counts = Counter(row.get("status", "") or "Unknown" for row in rows)
    priority_counts = Counter(row.get("priority", "") or "None" for row in rows)
    profile_counts = Counter(row.get("profile_id", "") or "Unknown" for row in rows)
    duplicate_count = sum(
        1
        for row in rows
        if row.get("decision") in {"Duplicate", "Already Reviewed"}
        or row.get("status") in {"Duplicate", "Already Reviewed"}
    )
    summary: list[dict[str, Any]] = [*metadata, {"metric": "Duplicate / already reviewed", "value": duplicate_count}]
    summary.extend({"metric": f"Decision: {key}", "value": value} for key, value in sorted(decision_counts.items()))
    summary.extend({"metric": f"Status: {key}", "value": value} for key, value in sorted(status_counts.items()))
    summary.extend({"metric": f"Priority: {key}", "value": value} for key, value in sorted(priority_counts.items()))
    summary.extend({"metric": f"Profile: {key}", "value": value} for key, value in sorted(profile_counts.items()))
    return summary


def _diagnostic_rows(capture_result: CaptureRunResult) -> list[dict[str, Any]]:
    diagnostics = capture_result.capture_diagnostics
    health = capture_result.capture_health
    return [
        {"field": "run_id", "value": capture_result.run_id},
        {"field": "status", "value": capture_result.status},
        {"field": "profile_id", "value": capture_result.profile_id},
        {"field": "capture_mode_used", "value": diagnostics.capture_mode_used},
        {"field": "input_size", "value": diagnostics.input_size},
        {"field": "candidate_cards_found", "value": diagnostics.candidate_cards_found},
        {"field": "cards_accepted", "value": diagnostics.cards_accepted},
        {"field": "cards_rejected", "value": diagnostics.cards_rejected},
        {"field": "capture_confidence", "value": diagnostics.capture_confidence},
        {"field": "browser_automation_enabled", "value": str(health.browser_automation_enabled)},
        {"field": "last_run_status", "value": health.last_run_status or ""},
        {"field": "capture_warnings", "value": _join(capture_result.warnings)},
        {"field": "diagnostic_warnings", "value": _join(diagnostics.warnings)},
        {"field": "rejection_reasons", "value": _join(diagnostics.rejection_reasons)},
        {"field": "source_url_notes", "value": _join(diagnostics.source_url_extraction_notes)},
    ]


def _explanation_rows(capture_result: CaptureRunResult) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in capture_result.results:
        job = result.parsed_job
        decision = result.decision
        rows.append(
            {
                "title": job.title if job else "",
                "company": job.company if job else "",
                "decision": "Duplicate" if result.duplicate_preview else decision.decision if decision else "",
                "triggered_rules": _join(decision.triggered_rules if decision else []),
                "matched_positive_keywords": _join(decision.matched_positive_keywords if decision else []),
                "matched_risk_keywords": _join(decision.matched_risk_keywords if decision else []),
                "hard_discard_reasons": _join(decision.reasons if decision and decision.decision == "Discard" else []),
                "warnings": _join(decision.warnings if decision else []),
                "missing_information": _join(decision.missing_information if decision else []),
            }
        )
    return rows


def _write_xlsx(path: Path, capture_result: CaptureRunResult, rows: list[dict[str, Any]], include_raw_text: bool) -> None:
    columns = BASE_COLUMNS + (["raw_text"] if include_raw_text else [])
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _append_table(summary_sheet, ["metric", "value"], _summary_rows(capture_result, rows))

    sheet_specs = [
        ("All Reviewed Jobs", rows),
        ("Apply Today", [row for row in rows if row.get("status") == "Apply Today" or row.get("decision") == "Apply"]),
        ("Manual Review", [row for row in rows if row.get("status") == "Manual Review" or row.get("decision") == "Manual Review"]),
        ("Waiting Follow Up", [row for row in rows if row.get("status") in {"Waiting", "Follow Up"}]),
        (
            "Duplicates Reviewed",
            [
                row
                for row in rows
                if row.get("status") in {"Duplicate", "Already Reviewed"}
                or row.get("decision") in {"Duplicate", "Already Reviewed"}
            ],
        ),
    ]
    for title, sheet_rows in sheet_specs:
        _append_table(workbook.create_sheet(title), columns, sheet_rows)

    _append_table(workbook.create_sheet("Decision Explanations"), EXPLANATION_COLUMNS, _explanation_rows(capture_result))
    _append_table(workbook.create_sheet("Capture Diagnostics"), ["field", "value"], _diagnostic_rows(capture_result))
    workbook.save(path)


def _flatten_history_entry(entry: HistoryJobEntry, include_raw_text: bool) -> dict[str, Any]:
    row: dict[str, Any] = {
        "status": entry.application_status,
        "decision": entry.decision,
        "priority": entry.priority,
        "score": entry.score,
        "title": entry.title,
        "company": entry.company,
        "location": entry.location,
        "work_mode": entry.work_mode,
        "profile_id": entry.profile_id,
        "parser_confidence": entry.parser_confidence,
        "capture_confidence": "",
        "reasons": _join(entry.reasons),
        "triggered_rules": "",
        "warnings": _join(entry.warnings),
        "missing_information": _join(entry.missing_information),
        "matched_positive_keywords": _join(entry.matched_positive_keywords),
        "matched_risk_keywords": _join(entry.matched_risk_keywords),
        "source_url": entry.source_url,
        "created_at": entry.saved_at,
        "errors": "",
    }
    if include_raw_text:
        row["raw_text"] = entry.raw_text or ""
    return row


def _history_explanation_rows(entries: list[HistoryJobEntry]) -> list[dict[str, Any]]:
    return [
        {
            "title": entry.title,
            "company": entry.company,
            "decision": entry.decision,
            "triggered_rules": "",
            "matched_positive_keywords": _join(entry.matched_positive_keywords),
            "matched_risk_keywords": _join(entry.matched_risk_keywords),
            "hard_discard_reasons": _join(entry.reasons if entry.decision == "Discard" else []),
            "warnings": _join(entry.warnings),
            "missing_information": _join(entry.missing_information),
        }
        for entry in entries
    ]


def _write_history_xlsx(
    path: Path,
    entries: list[HistoryJobEntry],
    rows: list[dict[str, Any]],
    include_raw_text: bool,
) -> None:
    columns = BASE_COLUMNS + (["raw_text"] if include_raw_text else [])
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _append_table(
        summary_sheet,
        ["metric", "value"],
        _summary_rows_for_export(
            rows,
            metadata=[
                {"metric": "Export timestamp", "value": datetime.now(UTC).isoformat(timespec="seconds")},
                {"metric": "Export source", "value": "Saved History / Tracker"},
                {"metric": "Capture mode", "value": "history_tracker"},
                {"metric": "Capture confidence", "value": ""},
                {"metric": "Total jobs", "value": len(rows)},
            ],
        ),
    )
    sheet_specs = [
        ("All Reviewed Jobs", rows),
        ("Apply Today", [row for row in rows if row.get("status") == "Apply Today" or row.get("decision") == "Apply"]),
        ("Manual Review", [row for row in rows if row.get("status") == "Manual Review" or row.get("decision") == "Manual Review"]),
        ("Waiting Follow Up", [row for row in rows if row.get("status") in {"Waiting", "Follow Up"}]),
        (
            "Duplicates Reviewed",
            [
                row
                for row in rows
                if row.get("status") in {"Duplicate", "Already Reviewed"}
                or row.get("decision") in {"Duplicate", "Already Reviewed"}
            ],
        ),
    ]
    for title, sheet_rows in sheet_specs:
        _append_table(workbook.create_sheet(title), columns, sheet_rows)

    _append_table(workbook.create_sheet("Decision Explanations"), EXPLANATION_COLUMNS, _history_explanation_rows(entries))
    _append_table(
        workbook.create_sheet("Capture Diagnostics"),
        ["field", "value"],
        [
            {"field": "export_source", "value": "history_tracker"},
            {"field": "history_rows", "value": len(entries)},
            {"field": "note", "value": "Tracker export uses latest saved statuses from local history."},
        ],
    )
    workbook.save(path)


def _write_history_json(path: Path, entries: list[HistoryJobEntry], include_raw_text: bool) -> None:
    payload = {
        "export_source": "history_tracker",
        "exported_at": datetime.now(UTC).isoformat(timespec="seconds"),
        "jobs": [entry.model_dump() for entry in entries],
    }
    if not include_raw_text:
        for job in payload["jobs"]:
            job["raw_text"] = None
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8",)


def export_capture_result(
    capture_result: CaptureRunResult,
    export_format: ExportFormat,
    include_raw_text: bool = False,
) -> ExportCaptureResultResponse:
    export_id = f"export_{uuid4().hex}"
    export_dir = _ensure_export_dir(export_id)
    rows = [_flatten_result(result, include_raw_text) for result in capture_result.results]
    for row in rows:
        row["capture_confidence"] = capture_result.capture_diagnostics.capture_confidence
    warnings: list[str] = []

    suffix = export_format
    path = export_dir / f"{capture_result.run_id}.{suffix}"
    if export_format == "json":
        _write_json(path, capture_result, include_raw_text)
    elif export_format == "csv":
        _write_csv(path, rows, include_raw_text)
    elif export_format == "xlsx":
        _write_xlsx(path, capture_result, rows, include_raw_text)
    else:  # pragma: no cover - Pydantic validates this before service entry.
        raise ValueError(f"Unsupported export format: {export_format}")

    if not include_raw_text:
        warnings.append("Raw job text was excluded from the export.")

    return ExportCaptureResultResponse(
        export_id=export_id,
        status="completed",
        files=[_relative_to_backend(path)],
        warnings=warnings,
    )


def export_history_entries(
    entries: list[HistoryJobEntry],
    export_format: ExportFormat,
    include_raw_text: bool = False,
) -> ExportCaptureResultResponse:
    export_id = f"export_{uuid4().hex}"
    export_dir = _ensure_export_dir(export_id)
    rows = [_flatten_history_entry(entry, include_raw_text) for entry in entries]
    path = export_dir / f"history_tracker.{export_format}"
    warnings: list[str] = []

    if export_format == "json":
        _write_history_json(path, entries, include_raw_text)
    elif export_format == "csv":
        _write_csv(path, rows, include_raw_text)
    elif export_format == "xlsx":
        _write_history_xlsx(path, entries, rows, include_raw_text)
    else:  # pragma: no cover - Pydantic validates this before service entry.
        raise ValueError(f"Unsupported export format: {export_format}")

    if not include_raw_text:
        warnings.append("Raw job text was excluded from the export.")
    warnings.append("Exported saved tracker/history data with latest statuses.")

    return ExportCaptureResultResponse(
        export_id=export_id,
        status="completed",
        files=[_relative_to_backend(path)],
        warnings=warnings,
    )
